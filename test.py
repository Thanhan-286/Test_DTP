from openpyxl import load_workbook
from datetime import time, timedelta

# Map trung tâm với thành phố nội thành
CENTER_CITY_MAP = {
    4: 'Thành phố Hà Nội',
    2: 'Thành phố Hồ Chí Minh',
    3: 'Thành phố Đà Nẵng',
}


def is_inner_city(city, center):
    # xác định đơn nội thành (True) hay ngoại thành (False)
    if city is None or center is None:
        return None

    try:
        center_int = int(center)
    except Exception:
        return None

    main_city = CENTER_CITY_MAP.get(center_int)
    if main_city is None:
        return None

    return city == main_city


def classify_inner_case(created_at):
    # phân loại đơn nội thành theo giờ đặt
    t = created_at.time()

    if t < time(10, 0):
        return 'inner_case1'
    elif t < time(16, 0):
        return 'inner_case2'
    else:
        return 'inner_case3'


def check_inner_case_sla(case_code, created_at, approved_at, packed_at, completed_at):
    created_date = created_at.date()

    if case_code == 'inner_case1':
        # trước 10h
        return (
            approved_at.date() == created_date
            and approved_at.time() <= time(12, 0)
            and packed_at.date() == created_date
            and packed_at.time() <= time(14, 0)
            and completed_at.date() == created_date
        )

    elif case_code == 'inner_case2':
        # 10h - 16h
        next_date = created_date + timedelta(days=1)
        return (
            approved_at.date() == created_date
            and approved_at.time() <= time(17, 0)
            and packed_at.date() == next_date
            and packed_at.time() <= time(10, 0)
            and completed_at.date() == next_date
            and completed_at.time() <= time(10, 0)
        )

    elif case_code == 'inner_case3':
        # sau 16h
        next_date = created_date + timedelta(days=1)
        return (
            approved_at.date() == next_date
            and approved_at.time() <= time(9, 30)
            and packed_at.date() == next_date
            and packed_at.time() <= time(12, 0)
            and completed_at.date() == next_date
            and completed_at.time() <= time(16, 0)
        )

    return False


def check_outer_sla(created_at, approved_at, packed_at, completed_at):
    # SLA đơn ngoại thành
    return (
        (approved_at - created_at) <= timedelta(hours=24)
        and (packed_at - approved_at) <= timedelta(hours=4)
        and (completed_at - packed_at) <= timedelta(hours=24)
    )


def analyze_file(filename, sheet_name='Sheet1'):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]

    stats = {
        'inner_case1': {'total': 0, 'on_time': 0},
        'inner_case2': {'total': 0, 'on_time': 0},
        'inner_case3': {'total': 0, 'on_time': 0},
        'outer_case': {'total': 0, 'on_time': 0},
    }
    stats_center_city = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        created_at, approved_at, packed_at, completed_at, city, center = row[:6]

        if (
            created_at is None or approved_at is None or
            packed_at is None or completed_at is None or
            city is None or center is None
        ):
            continue

        inner_flag = is_inner_city(city, center)
        if inner_flag is None:
            continue

        if inner_flag:
            # lọc 3 TH nội thành
            case_code = classify_inner_case(created_at)
            stats[case_code]['total'] += 1
            on_time = check_inner_case_sla(case_code, created_at, approved_at, packed_at, completed_at)
            if on_time:
                stats[case_code]['on_time'] += 1
        else:
            case_code = 'outer_case'
            stats[case_code]['total'] += 1
            on_time = check_outer_sla(created_at, approved_at, packed_at, completed_at)
            if on_time:
                stats[case_code]['on_time'] += 1

        # Thống kê theo (center, city)
        key = (center, city)
        if key not in stats_center_city:
            stats_center_city[key] = {'total': 0, 'on_time': 0}
        stats_center_city[key]['total'] += 1
        if on_time:
            stats_center_city[key]['on_time'] += 1

    case_names = {
        'inner_case1': 'Nội thành - Đặt trước 10h',
        'inner_case2': 'Nội thành - Đặt 10h-16h',
        'inner_case3': 'Nội thành - Đặt sau 16h',
        'outer_case': 'Ngoại thành',
    }

    for code, info in stats.items():
        total = info['total']
        # late = total - info['on_time']
        late = info['on_time']
        print(f"=== {case_names[code]} ===")
        print(f"  Tổng đơn : {total}")
        print(f"  Trễ SLA  : {late}")
        print()

    print("=== THỐNG KÊ THEO CENTER & CITY ===")
    for (center, city), info in stats_center_city.items():
        total = info['total']
        late = total - info['on_time']
        print(f"Center {center} - {city}: Tổng {total} đơn, Trễ SLA {late} đơn")

    return stats, stats_center_city


if __name__ == "__main__":
    analyze_file("Test.xlsx")
